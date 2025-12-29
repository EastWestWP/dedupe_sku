import sys
import re
from pathlib import Path
from decimal import Decimal, InvalidOperation
import pandas as pd
from openpyxl import load_workbook

# Update these if your spreadsheet uses different headers:
SKU_COL = "Variant SKU"
PRICE_COL = "Variant Price"
TITLE_COL = "Title"  # Output will be sorted by this column (like the original export)

# Columns that should ALWAYS be treated as text to preserve full length.
# By default, we will also auto-detect any column name containing "id".
FORCE_TEXT_COLS = set([
    "ID",
    "Variant ID",
    "Image ID",
    "Inventory Item ID",
])

SCI_RE = re.compile(r"^[+-]?\d+(\.\d+)?[eE][+-]?\d+$")


def die(msg: str, code: int = 1):
    print(msg)
    input("Press Enter to close...")
    raise SystemExit(code)


def to_plain_string(x) -> str:
    # Convert numeric/scientific-notation-like values to a plain string.
    # Keeps integers without .0, preserves leading zeros if already a string.
    if x is None:
        return ""

    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return ""

    # If it's already a non-scientific string, keep as-is
    if not SCI_RE.match(s):
        # Remove trailing .0 if it came from numeric conversion but looks integer
        if re.match(r"^\d+\.0$", s):
            return s[:-2]
        return s

    # Convert scientific notation to plain string
    try:
        d = Decimal(s)
        if d == d.to_integral_value():
            return format(d.quantize(Decimal(1)), "f")
        return format(d.normalize(), "f")
    except (InvalidOperation, ValueError):
        return s


def main():
    if len(sys.argv) < 2:
        die("Usage: python dedupe_sku_keep_max_price.py <input.xlsx>")

    input_path = Path(sys.argv[1]).expanduser()

    if not input_path.exists():
        die(f"File not found: {input_path}")

    if input_path.suffix.lower() not in [".xlsx", ".xlsm", ".xls"]:
        die("Input must be an Excel file (.xlsx/.xlsm/.xls)")

    try:
        # Read ALL columns as strings to avoid Excel/scientific notation issues.
        df = pd.read_excel(input_path, dtype=str)
    except Exception as e:
        die(f"Failed to read Excel: {e}")

    if SKU_COL not in df.columns:
        die(f"Missing column: '{SKU_COL}'")
    if PRICE_COL not in df.columns:
        die(f"Missing column: '{PRICE_COL}'")

    # Keep original row order for stable tie-breaking and sorting output like original export
    df["__row"] = range(len(df))

    # Identify columns that must be text (any column whose name contains "id", plus FORCE_TEXT_COLS)
    id_like_cols = [c for c in df.columns if "id" in str(c).lower()]
    force_text = set(id_like_cols) | set(FORCE_TEXT_COLS)

    # Normalize ID-like columns to plain strings (convert scientific notation, drop .0)
    for c in df.columns:
        if c in force_text:
            df[c] = df[c].map(to_plain_string)

    # Create numeric price for sorting, but keep original PRICE_COL string as-is
    price_num = pd.to_numeric(df[PRICE_COL].astype(str).str.replace(",", ""), errors="coerce")
    df["__price_num"] = price_num

    # Select the row with the highest price per SKU.
    # Tie-breaker: the earliest row in the original file.
    pick_sorted = df.sort_values([SKU_COL, "__price_num", "__row"], ascending=[True, False, True], kind="mergesort")
    deduped = pick_sorted.drop_duplicates(subset=[SKU_COL], keep="first").copy()

    removed = df.loc[~df["__row"].isin(deduped["__row"])].copy()

    # Sort output by Title (like original file). If Title is missing, keep original order.
    if TITLE_COL in deduped.columns:
        deduped = deduped.sort_values([TITLE_COL, "__row"], ascending=[True, True], kind="mergesort")
        removed = removed.sort_values([TITLE_COL, "__row"], ascending=[True, True], kind="mergesort")
    else:
        deduped = deduped.sort_values(["__row"], kind="mergesort")
        removed = removed.sort_values(["__row"], kind="mergesort")

    # Drop helper cols before saving
    deduped = deduped.drop(columns=["__price_num", "__row"])
    removed = removed.drop(columns=["__price_num", "__row"])

    out_base = input_path.with_suffix("")
    output_file = Path(str(out_base) + "_deduped.xlsx")
    removed_file = Path(str(out_base) + "_removed_duplicates.xlsx")

    try:
        deduped.to_excel(output_file, index=False)
        removed.to_excel(removed_file, index=False)
    except Exception as e:
        die(f"Failed to write Excel output: {e}")

    # Post-process Excel files to FORCE text formatting for ID-like columns
    def enforce_text_format(path: Path):
        wb = load_workbook(path)
        ws = wb.active

        # Map header -> column index
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=col_idx).value
            if v is not None:
                headers[str(v)] = col_idx

        target_cols = [headers[c] for c in headers.keys() if "id" in c.lower() or c in force_text]

        for col_idx in target_cols:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is None:
                    continue
                cell.number_format = "@"
                cell.value = to_plain_string(cell.value)

            ws.cell(row=1, column=col_idx).number_format = "@"

        wb.save(path)

    try:
        enforce_text_format(output_file)
        enforce_text_format(removed_file)
    except Exception as e:
        die(f"Wrote output, but failed to enforce text formatting: {e}")

    print("Done.")
    print(f"Saved cleaned file: {output_file}")
    print(f"Saved removed rows: {removed_file}")
    if TITLE_COL in df.columns:
        print(f"Sorted output by: {TITLE_COL}")
    else:
        print("Title column not found; kept original row order.")
    print("Note: All ID-like columns are saved as TEXT to avoid scientific notation.")
    input("Press Enter to close...")


if __name__ == "__main__":
    main()
